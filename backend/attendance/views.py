from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.db.models import Q, Count, Avg
from datetime import datetime, date, timedelta
from .models import Presence, Retard, Absence, BiometricLog
from .serializers import (
    PresenceSerializer, RetardSerializer, AbsenceSerializer, BiometricLogSerializer,
    RetardJustificationSerializer, RetardValidationSerializer,
    AbsenceJustificationSerializer, AbsenceValidationSerializer,
    BiometricLogCreateSerializer
)
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO

User = get_user_model()

class IsRHOrReadOnly(permissions.BasePermission):
    """
    Permission personnalisée : RH peut tout faire, autres utilisateurs lecture seule
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.role in ['DG', 'RH']

class IsOwnerOrRH(permissions.BasePermission):
    """
    Permission : propriétaire ou RH peut modifier
    """
    def has_object_permission(self, request, view, obj):
        if request.user.role in ['DG', 'RH']:
            return True
        return obj.employee == request.user

class IsEmployeOrRHOrReadOnly(permissions.BasePermission):
    """
    Permission : DG, RH et EMPLOYE peuvent faire POST (pour le pointage), autres lecture seule
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.role in ['DG', 'RH', 'EMPLOYE']

class PresenceViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des présences
    """
    serializer_class = PresenceSerializer
    permission_classes = [IsRHOrReadOnly]
    
    def get_queryset(self):
        """Filtrer les présences selon le rôle de l'utilisateur"""
        user = self.request.user
        
        if user.role in ['DG', 'RH']:
            # RH et DG voient toutes les présences
            queryset = Presence.objects.all()
        else:
            # Employé ne voit que ses propres présences
            queryset = Presence.objects.filter(employee=user)
        
        # Filtres optionnels
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        employee_id = self.request.query_params.get('employee_id')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if employee_id and user.role in ['DG', 'RH']:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.select_related('employee').order_by('-date', '-created_at')
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Obtenir les statistiques de présence"""
        user = request.user
        # Période par défaut : 30 derniers jours
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        # Ajuster selon les paramètres
        if 'start_date' in request.query_params:
            start_date = datetime.strptime(request.query_params['start_date'], '%Y-%m-%d').date()
        if 'end_date' in request.query_params:
            end_date = datetime.strptime(request.query_params['end_date'], '%Y-%m-%d').date()
        # Filtrer les présences
        if user.role in ['DG', 'RH']:
            presences = Presence.objects.filter(date__range=[start_date, end_date])
        else:
            presences = Presence.objects.filter(employee=user, date__range=[start_date, end_date])
        # Calculer les statistiques
        total_presences = presences.count()
        # Filtrer les absences selon le rôle
        if user.role in ['DG', 'RH']:
            total_absences = Absence.objects.filter(date__range=[start_date, end_date]).count()
            total_retards = Retard.objects.filter(date__range=[start_date, end_date]).count()
        else:
            total_absences = Absence.objects.filter(employee=user, date__range=[start_date, end_date]).count()
            total_retards = Retard.objects.filter(employee=user, date__range=[start_date, end_date]).count()
        # Heures moyennes travaillées (calcul Python)
        total_hours = 0
        count_days = 0
        for p in presences:
            if p.time_in and p.time_out:
                start = datetime.combine(p.date, p.time_in)
                end = datetime.combine(p.date, p.time_out)
                hours = (end - start).total_seconds() / 3600
                total_hours += hours
                count_days += 1
        avg_hours = round(total_hours / count_days, 2) if count_days > 0 else 0
        return Response({
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'statistics': {
                'total_presences': total_presences,
                'total_absences': total_absences,
                'total_retards': total_retards,
                'avg_hours_per_day': avg_hours
            }
        })
    
    @action(detail=False, methods=['post'])
    def manual_punch(self, request):
        """Pointage manuel (pour les tests ou corrections)"""
        if request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        employee_id = request.data.get('employee_id')
        punch_type = request.data.get('punch_type')  # 'in' ou 'out'
        punch_time = request.data.get('punch_time')  # 'HH:MM' ou None pour maintenant
        
        try:
            employee = User.objects.get(id=employee_id, is_active=True)
        except User.DoesNotExist:
            return Response(
                {'error': 'Employé non trouvé'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Heure de pointage
        if punch_time:
            punch_datetime = datetime.strptime(f"{date.today()} {punch_time}", "%Y-%m-%d %H:%M")
        else:
            punch_datetime = timezone.now()
        
        # Créer ou mettre à jour la présence
        presence, created = Presence.objects.get_or_create(
            employee=employee,
            date=punch_datetime.date(),
            defaults={}
        )
        
        if punch_type == 'in':
            presence.time_in = punch_datetime.time()
        elif punch_type == 'out':
            presence.time_out = punch_datetime.time()
        
        presence.save()
        
        serializer = self.get_serializer(presence)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='employee-punch', permission_classes=[IsEmployeOrRHOrReadOnly])
    def employee_punch(self, request):
        """
        Permet à un employé de pointer sa présence (simulateur biométrique)
        - L'heure du pointage est celle du serveur
        - Refuse si hors horaires (avant 6h ou après 19h)
        """
        user = request.user
        now = timezone.localtime()
        heure = now.time()
        # Heures autorisées : 06:00 <= heure < 19:00
        if not (heure >= datetime.strptime('06:00', '%H:%M').time() and heure < datetime.strptime('19:00', '%H:%M').time()):
            return Response({
                'success': False,
                'message': "Pointage refusé : hors horaires autorisés (6h00 à 19h00)."
            }, status=status.HTTP_403_FORBIDDEN)
        # Créer ou mettre à jour la présence du jour
        presence, created = Presence.objects.get_or_create(
            employee=user,
            date=now.date(),
            defaults={}
        )
        if presence.time_in:
            return Response({
                'success': False,
                'message': "Vous avez déjà pointé aujourd'hui."
            }, status=status.HTTP_400_BAD_REQUEST)
        presence.time_in = heure
        presence.save()
        return Response({
            'success': True,
            'message': "Présence enregistrée avec succès !",
            'heure': heure.strftime('%H:%M')
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='employee-punch-out', permission_classes=[IsEmployeOrRHOrReadOnly])
    def employee_punch_out(self, request):
        """
        Permet à un employé de pointer son départ (simulateur biométrique)
        - L'heure du pointage est celle du serveur
        - Autorisé uniquement si heure >= 18h00 ou < 6h00
        """
        user = request.user
        now = timezone.localtime()
        heure = now.time()
        # Chercher la présence du jour
        try:
            presence = Presence.objects.get(employee=user, date=now.date())
        except Presence.DoesNotExist:
            return Response({
                'success': False,
                'message': "Vous devez d'abord pointer votre arrivée."
            }, status=status.HTTP_400_BAD_REQUEST)
        if presence.time_out:
            return Response({
                'success': False,
                'message': "Vous avez déjà pointé votre départ aujourd'hui."
            }, status=status.HTTP_400_BAD_REQUEST)
        # Autorisé si heure >= 18h00 ou < 6h00
        if not (heure >= datetime.strptime('18:00', '%H:%M').time() or heure < datetime.strptime('06:00', '%H:%M').time()):
            return Response({
                'success': False,
                'message': "Pointage de départ refusé : vous ne pouvez pointer votre départ qu'à partir de 18h00 ou avant 6h00 le lendemain."
            }, status=status.HTTP_403_FORBIDDEN)
        presence.time_out = heure
        presence.save()
        return Response({
            'success': True,
            'message': "Départ enregistré avec succès !",
            'heure': heure.strftime('%H:%M')
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export')
    def export_pdf(self, request):
        """
        Génère un PDF de l'historique des présences pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé. Ajoute un résumé clair en haut du PDF.
        """
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        # Préparation du PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="presences.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - 2*cm
        p.setFont("Helvetica-Bold", 16)
        p.drawString(2*cm, y, "Rapport de présences")
        y -= 1*cm
        p.setFont("Helvetica", 10)
        p.drawString(2*cm, y, f"Employé : {user.get_full_name()} ({user.email})")
        y -= 0.7*cm
        # Affichage de la période
        if date_from and date_to:
            p.drawString(2*cm, y, f"Période : du {date_from} au {date_to}")
        elif date_from:
            p.drawString(2*cm, y, f"Période : à partir du {date_from}")
        elif date_to:
            p.drawString(2*cm, y, f"Période : jusqu'au {date_to}")
        else:
            p.drawString(2*cm, y, "Période : toutes les données disponibles")
        y -= 0.7*cm
        p.drawString(2*cm, y, f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 0.7*cm
        # Note explicative
        p.setFont("Helvetica-Oblique", 9)
        p.drawString(2*cm, y, "Ce rapport présente l'ensemble de vos présences enregistrées sur la période sélectionnée. Pour toute question, contactez le service RH.")
        y -= 1*cm
        # En-tête du tableau
        p.setFont("Helvetica-Bold", 10)
        p.drawString(2*cm, y, "Date")
        p.drawString(5*cm, y, "Entrée")
        p.drawString(8*cm, y, "Sortie")
        p.drawString(11*cm, y, "Retard")
        p.drawString(14*cm, y, "Total h")
        p.drawString(17*cm, y, "Statut")
        y -= 0.5*cm
        p.setFont("Helvetica", 10)
        for presence in queryset:
            if y < 2*cm:
                p.showPage()
                y = height - 2*cm
            p.drawString(2*cm, y, presence.date.strftime('%d/%m/%Y'))
            p.drawString(5*cm, y, presence.time_in.strftime('%H:%M') if presence.time_in else '-')
            p.drawString(8*cm, y, presence.time_out.strftime('%H:%M') if presence.time_out else '-')
            p.drawString(11*cm, y, f"{presence.delay_minutes} min" if presence.is_late else '-')
            # Calcul total heures
            if presence.time_in and presence.time_out:
                start = datetime.combine(presence.date, presence.time_in)
                end = datetime.combine(presence.date, presence.time_out)
                hours = round((end - start).total_seconds() / 3600, 2)
                p.drawString(14*cm, y, f"{hours}h")
            else:
                p.drawString(14*cm, y, '-')
            # Statut
            if not presence.time_in:
                statut = 'ABSENT'
            elif not presence.time_out:
                statut = 'EN COURS'
            else:
                statut = 'TERMINÉ'
            p.drawString(17*cm, y, statut)
            y -= 0.5*cm
        p.showPage()
        p.save()
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Génère un fichier Excel de l'historique des présences pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé.
        """
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Présences"
        # Titre
        ws.merge_cells('A1:G1')
        ws['A1'] = "Rapport de présences"
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        ws.append([f"Employé : {user.get_full_name()} ({user.email})"])
        if date_from and date_to:
            ws.append([f"Période : du {date_from} au {date_to}"])
        elif date_from:
            ws.append([f"Période : à partir du {date_from}"])
        elif date_to:
            ws.append([f"Période : jusqu'au {date_to}"])
        else:
            ws.append(["Période : toutes les données disponibles"])
        from datetime import datetime
        ws.append([f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append(["Ce rapport présente l'ensemble de vos présences enregistrées sur la période sélectionnée. Pour toute question, contactez le service RH."])
        ws.append([])
        # En-tête
        headers = ["Date", "Entrée", "Sortie", "Retard", "Total h", "Statut"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        # Données
        for presence in queryset:
            date_str = presence.date.strftime('%d/%m/%Y')
            entree = presence.time_in.strftime('%H:%M') if presence.time_in else '-'
            sortie = presence.time_out.strftime('%H:%M') if presence.time_out else '-'
            retard = f"{presence.delay_minutes} min" if presence.is_late else '-'
            if presence.time_in and presence.time_out:
                start = datetime.combine(presence.date, presence.time_in)
                end = datetime.combine(presence.date, presence.time_out)
                total_h = round((end - start).total_seconds() / 3600, 2)
                total_h_str = f"{total_h}h"
            else:
                total_h_str = '-'
            if not presence.time_in:
                statut = 'ABSENT'
            elif not presence.time_out:
                statut = 'EN COURS'
            else:
                statut = 'TERMINÉ'
            ws.append([date_str, entree, sortie, retard, total_h_str, statut])
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        # Sauvegarder dans un buffer mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="presences.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='rh-dashboard')
    def rh_dashboard(self, request):
        """
        Dashboard synthétique RH : nombre d'employés, absences/retards en attente, stats globales
        """
        user = request.user
        if user.role not in ['DG', 'RH']:
            return Response({'error': 'Permission refusée'}, status=status.HTTP_403_FORBIDDEN)
        from users.models import User
        total_employees = User.objects.filter(is_active=True, role='EMPLOYE').count()
        absences_en_attente = Absence.objects.filter(justification_status='EN_ATTENTE').count()
        retards_en_attente = Retard.objects.filter(justification_status='EN_ATTENTE').count()
        # Statistiques globales (30 derniers jours)
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        presences = Presence.objects.filter(date__range=[start_date, end_date])
        total_presences = presences.count()
        total_absences = Absence.objects.filter(date__range=[start_date, end_date]).count()
        total_retards = Retard.objects.filter(date__range=[start_date, end_date]).count()
        # Heures moyennes travaillées
        total_hours = 0
        count_days = 0
        for p in presences:
            if p.time_in and p.time_out:
                start = datetime.combine(p.date, p.time_in)
                end = datetime.combine(p.date, p.time_out)
                hours = (end - start).total_seconds() / 3600
                total_hours += hours
                count_days += 1
        avg_hours = round(total_hours / count_days, 2) if count_days > 0 else 0
        return Response({
            'total_employees': total_employees,
            'absences_en_attente': absences_en_attente,
            'retards_en_attente': retards_en_attente,
            'statistics': {
                'total_presences': total_presences,
                'total_absences': total_absences,
                'total_retards': total_retards,
                'avg_hours_per_day': avg_hours
            }
        })

class RetardViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet pour la gestion des retards
    """
    serializer_class = RetardSerializer
    permission_classes = [IsRHOrReadOnly]
    
    def get_queryset(self):
        """Filtrer les retards selon le rôle de l'utilisateur"""
        user = self.request.user
        
        if user.role in ['DG', 'RH']:
            queryset = Retard.objects.all()
        else:
            queryset = Retard.objects.filter(employee=user)
        
        # Filtres
        status_filter = self.request.query_params.get('status')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        employee_id = self.request.query_params.get('employee_id')
        
        if status_filter:
            queryset = queryset.filter(justification_status=status_filter)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if employee_id and user.role in ['DG', 'RH']:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.select_related('employee', 'presence', 'validated_by').order_by('-date', '-created_at')
    
    @action(detail=True, methods=['patch'])
    def justify(self, request, pk=None):
        """Justifier un retard"""
        retard = self.get_object()
        
        # Vérifier que l'utilisateur peut justifier ce retard
        if request.user != retard.employee and request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = RetardJustificationSerializer(
            retard, 
            data=request.data, 
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['patch'])
    def validate(self, request, pk=None):
        """Valider ou refuser une justification de retard (RH uniquement)"""
        if request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        retard = self.get_object()
        serializer = RetardValidationSerializer(
            retard, 
            data=request.data, 
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='export')
    def export_pdf(self, request):
        """
        Génère un PDF de l'historique des retards pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé. Ajoute un résumé clair en haut du PDF.
        """
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="retards.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - 2*cm
        p.setFont("Helvetica-Bold", 16)
        p.drawString(2*cm, y, "Rapport de retards")
        y -= 1*cm
        p.setFont("Helvetica", 10)
        p.drawString(2*cm, y, f"Employé : {user.get_full_name()} ({user.email})")
        y -= 0.7*cm
        if date_from and date_to:
            p.drawString(2*cm, y, f"Période : du {date_from} au {date_to}")
        elif date_from:
            p.drawString(2*cm, y, f"Période : à partir du {date_from}")
        elif date_to:
            p.drawString(2*cm, y, f"Période : jusqu'au {date_to}")
        else:
            p.drawString(2*cm, y, "Période : toutes les données disponibles")
        y -= 0.7*cm
        p.drawString(2*cm, y, f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 0.7*cm
        p.setFont("Helvetica-Oblique", 9)
        p.drawString(2*cm, y, "Ce rapport présente l'ensemble de vos retards enregistrés sur la période sélectionnée. Pour toute question, contactez le service RH.")
        y -= 1*cm
        p.setFont("Helvetica-Bold", 10)
        p.drawString(2*cm, y, "Date")
        p.drawString(5*cm, y, "Heure prévue")
        p.drawString(8*cm, y, "Heure réelle")
        p.drawString(11*cm, y, "Retard")
        p.drawString(14*cm, y, "Justification")
        p.drawString(18*cm, y, "Statut")
        y -= 0.5*cm
        p.setFont("Helvetica", 10)
        for retard in queryset:
            if y < 2*cm:
                p.showPage()
                y = height - 2*cm
            p.drawString(2*cm, y, retard.date.strftime('%d/%m/%Y'))
            p.drawString(5*cm, y, retard.expected_time.strftime('%H:%M'))
            p.drawString(8*cm, y, retard.actual_time.strftime('%H:%M'))
            p.drawString(11*cm, y, f"{retard.delay_minutes} min")
            justification = (retard.justification[:30] + '...') if retard.justification and len(retard.justification) > 30 else (retard.justification or '-')
            p.drawString(14*cm, y, justification)
            statut = dict(retard.STATUS_CHOICES).get(retard.justification_status, retard.justification_status)
            p.drawString(18*cm, y, statut)
            y -= 0.5*cm
        p.showPage()
        p.save()
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Génère un fichier Excel de l'historique des retards pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé.
        """
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Retards"
        ws.merge_cells('A1:G1')
        ws['A1'] = "Rapport de retards"
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        ws.append([f"Employé : {user.get_full_name()} ({user.email})"])
        if date_from and date_to:
            ws.append([f"Période : du {date_from} au {date_to}"])
        elif date_from:
            ws.append([f"Période : à partir du {date_from}"])
        elif date_to:
            ws.append([f"Période : jusqu'au {date_to}"])
        else:
            ws.append(["Période : toutes les données disponibles"])
        from datetime import datetime
        ws.append([f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append(["Ce rapport présente l'ensemble de vos retards enregistrés sur la période sélectionnée. Pour toute question, contactez le service RH."])
        ws.append([])
        headers = ["Date", "Heure prévue", "Heure réelle", "Retard", "Justification", "Statut"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for retard in queryset:
            date_str = retard.date.strftime('%d/%m/%Y')
            heure_prev = retard.expected_time.strftime('%H:%M')
            heure_reelle = retard.actual_time.strftime('%H:%M')
            retard_str = f"{retard.delay_minutes} min"
            justification = (retard.justification[:30] + '...') if retard.justification and len(retard.justification) > 30 else (retard.justification or '-')
            statut = dict(retard.STATUS_CHOICES).get(retard.justification_status, retard.justification_status)
            ws.append([date_str, heure_prev, heure_reelle, retard_str, justification, statut])
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="retards.xlsx"'
        return response

class AbsenceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet pour la gestion des absences
    """
    serializer_class = AbsenceSerializer
    permission_classes = [IsRHOrReadOnly]
    
    def get_queryset(self):
        """Filtrer les absences selon le rôle de l'utilisateur"""
        user = self.request.user
        
        if user.role in ['DG', 'RH']:
            queryset = Absence.objects.all()
        else:
            queryset = Absence.objects.filter(employee=user)
        
        # Filtres
        status_filter = self.request.query_params.get('status')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        employee_id = self.request.query_params.get('employee_id')
        
        if status_filter:
            queryset = queryset.filter(justification_status=status_filter)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if employee_id and user.role in ['DG', 'RH']:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.select_related('employee', 'validated_by').order_by('-date', '-created_at')
    
    @action(detail=True, methods=['patch'])
    def justify(self, request, pk=None):
        """Justifier une absence"""
        absence = self.get_object()
        
        # Vérifier que l'utilisateur peut justifier cette absence
        if request.user != absence.employee and request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = AbsenceJustificationSerializer(
            absence, 
            data=request.data, 
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['patch'])
    def validate(self, request, pk=None):
        """Valider ou refuser une justification d'absence (RH uniquement)"""
        if request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        absence = self.get_object()
        serializer = AbsenceValidationSerializer(
            absence, 
            data=request.data, 
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='export')
    def export_pdf(self, request):
        """
        Génère un PDF de l'historique des absences pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé. Ajoute un résumé clair en haut du PDF.
        """
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="absences.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - 2*cm
        p.setFont("Helvetica-Bold", 16)
        p.drawString(2*cm, y, "Rapport d'absences")
        y -= 1*cm
        p.setFont("Helvetica", 10)
        p.drawString(2*cm, y, f"Employé : {user.get_full_name()} ({user.email})")
        y -= 0.7*cm
        if date_from and date_to:
            p.drawString(2*cm, y, f"Période : du {date_from} au {date_to}")
        elif date_from:
            p.drawString(2*cm, y, f"Période : à partir du {date_from}")
        elif date_to:
            p.drawString(2*cm, y, f"Période : jusqu'au {date_to}")
        else:
            p.drawString(2*cm, y, "Période : toutes les données disponibles")
        y -= 0.7*cm
        p.drawString(2*cm, y, f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 0.7*cm
        p.setFont("Helvetica-Oblique", 9)
        p.drawString(2*cm, y, "Ce rapport présente l'ensemble de vos absences enregistrées sur la période sélectionnée. Pour toute question, contactez le service RH.")
        y -= 1*cm
        p.setFont("Helvetica-Bold", 10)
        p.drawString(2*cm, y, "Date")
        p.drawString(5*cm, y, "Justification")
        p.drawString(12*cm, y, "Statut")
        p.drawString(16*cm, y, "Validé par")
        y -= 0.5*cm
        p.setFont("Helvetica", 10)
        for absence in queryset:
            if y < 2*cm:
                p.showPage()
                y = height - 2*cm
            p.drawString(2*cm, y, absence.date.strftime('%d/%m/%Y'))
            justification = (absence.justification[:40] + '...') if absence.justification and len(absence.justification) > 40 else (absence.justification or '-')
            p.drawString(5*cm, y, justification)
            statut = dict(absence.STATUS_CHOICES).get(absence.justification_status, absence.justification_status)
            p.drawString(12*cm, y, statut)
            p.drawString(16*cm, y, absence.validated_by.get_full_name() if absence.validated_by else '-')
            y -= 0.5*cm
        p.showPage()
        p.save()
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Génère un fichier Excel de l'historique des absences pour l'utilisateur connecté (ou tous si RH/DG), filtré par période si précisé.
        """
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        user = request.user
        queryset = self.get_queryset()
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Absences"
        ws.merge_cells('A1:F1')
        ws['A1'] = "Rapport d'absences"
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        ws.append([f"Employé : {user.get_full_name()} ({user.email})"])
        if date_from and date_to:
            ws.append([f"Période : du {date_from} au {date_to}"])
        elif date_from:
            ws.append([f"Période : à partir du {date_from}"])
        elif date_to:
            ws.append([f"Période : jusqu'au {date_to}"])
        else:
            ws.append(["Période : toutes les données disponibles"])
        from datetime import datetime
        ws.append([f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append(["Ce rapport présente l'ensemble de vos absences enregistrées sur la période sélectionnée. Pour toute question, contactez le service RH."])
        ws.append([])
        headers = ["Date", "Justification", "Statut", "Validé par"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for absence in queryset:
            date_str = absence.date.strftime('%d/%m/%Y')
            justification = (absence.justification[:40] + '...') if absence.justification and len(absence.justification) > 40 else (absence.justification or '-')
            statut = dict(absence.STATUS_CHOICES).get(absence.justification_status, absence.justification_status)
            valideur = absence.validated_by.get_full_name() if absence.validated_by else '-'
            ws.append([date_str, justification, statut, valideur])
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="absences.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='mes-absences')
    def mes_absences(self, request):
        """
        Retourne la liste des absences ET des retards de l'utilisateur connecté
        """
        user = request.user

        absences = Absence.objects.filter(employee=user).order_by('-date')
        retards = Retard.objects.filter(employee=user).order_by('-date')

        absences_data = AbsenceSerializer(absences, many=True).data
        retards_data = RetardSerializer(retards, many=True).data

        # On ajoute un champ 'type' pour différencier dans le front
        for a in absences_data:
            a['type'] = 'ABSENCE'
        for r in retards_data:
            r['type'] = 'RETARD'

        # Fusionne et trie par date (optionnel)
        all_data = absences_data + retards_data
        all_data.sort(key=lambda x: x['date'], reverse=True)

        return Response(all_data)

class BiometricLogViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des logs biométriques
    """
    serializer_class = BiometricLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filtrer les logs selon le rôle de l'utilisateur"""
        user = self.request.user
        
        if user.role in ['DG', 'RH']:
            queryset = BiometricLog.objects.all()
        else:
            queryset = BiometricLog.objects.filter(employee=user)
        
        return queryset.select_related('employee').order_by('-timestamp')
    
    def get_serializer_class(self):
        """Utiliser un sérialiseur différent pour la création via API"""
        if self.action == 'create':
            return BiometricLogCreateSerializer
        return BiometricLogSerializer
    
    @action(detail=False, methods=['post'])
    def receive_punch(self, request):
        """
        Endpoint pour recevoir les données de pointage du dispositif biométrique
        Format attendu :
        {
            "biometric_id": "12345",
            "log_type": "ENTREE",
            "timestamp": "2024-01-15T08:30:00Z",
            "device_id": "DEVICE_001",
            "raw_data": {...}
        }
        """
        serializer = BiometricLogCreateSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                log = serializer.save()
                return Response({
                    'success': True,
                    'message': 'Pointage enregistré avec succès',
                    'log_id': log.id,
                    'employee': log.employee.get_full_name() if log.employee else None
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def create_absences(self, request):
        """
        Créer automatiquement les absences pour les employés qui n'ont pas pointé
        Cette action peut être appelée manuellement ou via un cron job
        """
        if request.user.role not in ['DG', 'RH']:
            return Response(
                {'error': 'Permission refusée'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Date pour laquelle créer les absences (par défaut hier)
        absence_date = request.data.get('date', (date.today() - timedelta(days=1)).isoformat())
        
        # Récupérer tous les employés actifs
        employees = User.objects.filter(is_active=True, role='EMPLOYE')
        
        absences_created = 0
        
        for employee in employees:
            # Vérifier s'il y a une présence pour cette date
            presence_exists = Presence.objects.filter(
                employee=employee, 
                date=absence_date
            ).exists()
            
            # Vérifier s'il y a déjà une absence pour cette date
            absence_exists = Absence.objects.filter(
                employee=employee, 
                date=absence_date
            ).exists()
            
            # Créer l'absence si pas de présence et pas d'absence
            if not presence_exists and not absence_exists:
                Absence.objects.create(
                    employee=employee,
                    date=absence_date
                )
                absences_created += 1
        
        return Response({
            'success': True,
            'message': f'{absences_created} absences créées pour le {absence_date}',
            'absences_created': absences_created
        })